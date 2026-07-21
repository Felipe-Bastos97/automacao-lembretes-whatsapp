"""Avisos de vencimento pelo WhatsApp Web com modo de simulação seguro."""

from __future__ import annotations

import argparse
import csv
import logging
import re
import time
import webbrowser
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote

import openpyxl


PASTA_PROJETO = Path(__file__).resolve().parent
PLANILHA_PADRAO = PASTA_PROJETO / "outputs" / "exemplo" / "clientes_exemplo.xlsx"
ARQUIVO_ERROS = PASTA_PROJETO / "erros_cobrancas.csv"
COLUNAS_OBRIGATORIAS = ("cliente", "telefone", "vencimento", "link_pagamento")


@dataclass(frozen=True)
class Cobranca:
    cliente: str
    telefone: str
    vencimento: date
    link_pagamento: str


def somente_digitos(valor: object) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def converter_data(valor: object) -> date:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return datetime.strptime(str(valor).strip(), "%d/%m/%Y").date()


def carregar_cobrancas(caminho: Path) -> list[Cobranca]:
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if "Cobrancas" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("A planilha deve conter a aba 'Cobrancas'.")

    planilha = workbook["Cobrancas"]
    cabecalhos = [str(celula.value or "").strip().lower() for celula in planilha[1]]
    if tuple(cabecalhos[:4]) != COLUNAS_OBRIGATORIAS:
        workbook.close()
        raise ValueError("Cabeçalhos inválidos. Use: " + ", ".join(COLUNAS_OBRIGATORIAS))

    cobrancas: list[Cobranca] = []
    for numero_linha, valores in enumerate(planilha.iter_rows(min_row=2, values_only=True), 2):
        if not any(valores[:4]):
            continue
        try:
            telefone = somente_digitos(valores[1])
            if len(telefone) < 12:
                raise ValueError("telefone deve incluir código do país e DDD")
            link = str(valores[3] or "").strip()
            if not link.startswith("https://"):
                raise ValueError("link de pagamento deve começar com https://")
            cobrancas.append(
                Cobranca(
                    cliente=str(valores[0]).strip(),
                    telefone=telefone,
                    vencimento=converter_data(valores[2]),
                    link_pagamento=link,
                )
            )
        except (TypeError, ValueError) as erro:
            logging.warning("Linha %s ignorada: %s", numero_linha, erro)

    workbook.close()
    return cobrancas


def criar_mensagem(cobranca: Cobranca) -> str:
    return (
        f"Olá {cobranca.cliente}, tudo bem? Este é um lembrete de que seu "
        f"pagamento vence em {cobranca.vencimento:%d/%m/%Y}.\n"
        f"Link para pagamento: {cobranca.link_pagamento}\n"
        "Se o pagamento já foi realizado, desconsidere esta mensagem."
    )


def criar_link(cobranca: Cobranca) -> str:
    return (
        "https://web.whatsapp.com/send?"
        f"phone={cobranca.telefone}&text={quote(criar_mensagem(cobranca))}"
    )


def registrar_erro(cobranca: Cobranca, motivo: str) -> None:
    novo_arquivo = not ARQUIVO_ERROS.exists()
    with ARQUIVO_ERROS.open("a", newline="", encoding="utf-8") as arquivo:
        writer = csv.writer(arquivo)
        if novo_arquivo:
            writer.writerow(("cliente", "telefone", "motivo"))
        writer.writerow((cobranca.cliente, cobranca.telefone, motivo))


def processar(cobrancas: list[Cobranca], enviar: bool, intervalo: int) -> None:
    if enviar:
        try:
            import pyautogui
        except ImportError as erro:
            raise RuntimeError("Instale as dependências antes de usar --enviar") from erro

    for cobranca in cobrancas:
        link = criar_link(cobranca)
        if not enviar:
            print(f"[SIMULAÇÃO] {cobranca.cliente}: {link}")
            continue
        try:
            webbrowser.open(link)
            time.sleep(intervalo)
            pyautogui.press("enter")
            time.sleep(3)
            pyautogui.hotkey("ctrl", "w")
        except Exception as erro:
            logging.exception("Falha no envio para %s", cobranca.cliente)
            registrar_erro(cobranca, str(erro))


def argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--planilha", type=Path, default=PLANILHA_PADRAO)
    parser.add_argument("--enviar", action="store_true")
    parser.add_argument("--intervalo", type=int, default=12)
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    opcoes = argumentos()
    cobrancas = carregar_cobrancas(opcoes.planilha)
    logging.info("%s cobrança(s) válida(s) carregada(s)", len(cobrancas))
    processar(cobrancas, opcoes.enviar, opcoes.intervalo)


if __name__ == "__main__":
    main()
